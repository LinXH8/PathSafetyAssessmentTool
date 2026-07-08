"""Extract the CycleRAP v2.14 model + STM treatment catalog into runtime JSON.

Reads `backend/CycleRAP - model - generation v2.14-to supliers.xlsm` and emits
two CHECKED-IN files consumed by the backend at import time (the workbook is
never in the runtime path):

  backend/app/services/data/cyclerap_v214_model.json
      - lookup_tables   : risk multiplier + conditional flag per attribute option
                          (keyed by the module keys used in cyclerap_scoring.py)
      - facility_type   : DN3:DT8 multi-column table (bp_cond/vb_cond/vb_sev/vb_cf)
      - aadt_table      : DI3:DK19 breakpoints (exact cached floats)
      - speed_table     : DD93:DF243 per-integer-speed factors, km/h + mph columns
      - bands           : inclusive upper thresholds per crash type
      - baselines       : safest coded option per attribute (module keys)

  backend/app/services/data/stm_v214_treatments.json
      - 30 treatments {id, name, effects, trigger_sets[...]}
        trigger set = {"attrs": {field: [cats]}, "aadt": ranges|None,
                       "speed": ranges|None}
        Applicability: ANY set passes; within a set every attr group must match
        one of its listed categories AND the AADT/speed ranges must hold
        (STM rows 5/6/7/8/9 logic).

Known workbook quirks handled here (documented in the audit section of the
plan / docs):
  - STM outcome cells for T3 and T5 "Facility access" evaluate to 2
    (Inadequate) because their formulas point at 'Upload spec'!G26 instead of
    G25, contradicting the adjacent description column ("Adequate"). We follow
    the description (code 1).
  - The STM per-row match flags (col W) contain off-by-N references for a few
    rows (e.g. W46 '=U46=V44'); the intended semantics — input value equals the
    row's own category — are implemented instead.
  - 'Unaccessible facility - Part 2' (CycleRAP Model DB270) always evaluates to
    the facility's own category (OFFSET column self-match), so the CM42 trigger
    reduces to the facility-type VB-conditional column (DR). Verified against
    cached values.

Run:  backend/venv/bin/python scripts/extract_v214_model.py
"""

import json
import re
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[1]
XLSM = ROOT / "backend" / "CycleRAP - model - generation v2.14-to supliers.xlsm"
DATA_DIR = ROOT / "backend" / "app" / "services" / "data"
MODEL_OUT = DATA_DIR / "cyclerap_v214_model.json"
TREATMENTS_OUT = DATA_DIR / "stm_v214_treatments.json"

MODEL_SHEET = "CycleRAP Model"
STM_SHEET = "STM"

# --- column indices (1-based) -------------------------------------------------
from openpyxl.utils import column_index_from_string as ci


def norm_label(s):
    """Normalise a workbook attribute/option label for matching."""
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


# Workbook attribute label -> module key used in cyclerap_scoring.LOOKUP_TABLES
LABEL_TO_MODULE_KEY = {
    "facility access": "facility_access",
    "loose or slippery surface": "loose_surface",
    "tram or train rails": "tram_rails",
    "major surface deformation or drain opening": "surface_deformation",
    "fixed obstacle on facility": "fixed_obstacle",
    "non-fixed obstacle on facility": "non_fixed_obstacle",
    "line of sight": "line_of_sight",
    "sight distance": "line_of_sight",
    "delineation": "delineation",
    "light segregation": "light_segregation",
    "facility width per direction": "facility_width",
    "flow direction": "flow_direction",
    "width restriction": "width_restriction",
    "adjacent road lane 0-1m": "adjacent_road_0_1m",
    "adjacent vehicle parking 0-1m": "adjacent_parking_0_1m",
    "adjacent severe hazard 0-1m": "adjacent_hazard_0_1m",
    "adjacent object or level change 0-1m": "adjacent_object_0_1m",
    "adjacent sidewalk 0-1m": "adjacent_sidewalk_0_1m",
    "adjacent road lane 1-3m": "adjacent_road_1_3m",
    "adjacent vehicle parking 1-3m": "adjacent_parking_1_3m",
    "adjacent severe hazard 1-3m": "adjacent_hazard_1_3m",
    "adjacent object or level change 1-3m": "adjacent_object_1_3m",
    "adjacent sidewalk 1-3m": "adjacent_sidewalk_1_3m",
    "grade": "grade",
    "curvature": "curvature",
    "street lighting": "street_lighting",
    "pedestrian crossing": "pedestrian_crossing",
    "intersecting bicycle facility": "intersecting_facility",
    "intersection approach": "intersection_approach",
    "intersection or road crossing": "intersection_crossing",
    "crossing facility": "crossing_facility",
    "number of lanes – adjacent road": "num_lanes_adjacent",
    "number of lanes - adjacent road": "num_lanes_adjacent",
    "number of lanes – intersecting road": "num_lanes_intersecting",
    "number of lanes - intersecting road": "num_lanes_intersecting",
    "property access": "property_access",
    "peak pedestrian flow along or across facility": "pedestrian_flow",
    "peak bicycle/lv traffic flow": "bicycle_flow",
    "observed proportion of cargo bikes and mopeds": "cargo_bikes",
    "bicycle/lv speed – average": "bicycle_speed",
    "bicycle/lv speed - average": "bicycle_speed",
    "bicycle/lv speed differential": "speed_differential",
    "heavy vehicle flow": "heavy_vehicle",
}

# Workbook attribute label -> backend attributes.csv column name
# (Attributes.Fields.* in backend/app/services/serializer.py)
LABEL_TO_BACKEND_FIELD = {
    "area type": "Area type",
    "facility type": "Facility Type",
    "facility access": "Facility access",
    "line of sight": "Line of Sight",
    "sight distance": "Line of Sight",
    "loose or slippery surface": "Loose or slippery surface",
    "tram or train rails": "Tram or Train Rails",
    "major surface deformation or drain opening": "Major Surface Deformation or Drain Opening",
    "fixed obstacle on facility": "Fixed Obstacle on Facility",
    "non-fixed obstacle on facility": "Non-Fixed Obstacle on Facility",
    "delineation": "Delineation",
    "light segregation": "Light Segregation",
    "facility width per direction": "Facility Width per Direction",
    "flow direction": "Flow Direction",
    "width restriction": "Width Restriction",
    "adjacent road lane 0-1m": "Adjacent Road Lane 0-1m",
    "adjacent vehicle parking 0-1m": "Adjacent Vehicle Parking 0-1m",
    "adjacent severe hazard 0-1m": "Adjacent Severe Hazard 0-1m",
    "adjacent object or level change 0-1m": "Adjacent object or level change 0-1m",
    "adjacent sidewalk 0-1m": "Adjacent Sidewalk 0-1m",
    "adjacent road lane 1-3m": "Adjacent Road Lane 1-3m",
    "adjacent vehicle parking 1-3m": "Adjacent Vehicle Parking 1-3m",
    "adjacent severe hazard 1-3m": "Adjacent Severe Hazard 1-3m",
    "adjacent object or level change 1-3m": "Adjacent object or level change 1-3m",
    "adjacent sidewalk 1-3m": "Adjacent Sidewalk 1-3m",
    "grade": "Grade",
    "curvature": "Curvature",
    "street lighting": "Street Lighting",
    "pedestrian crossing": "Pedestrian Crossing",
    "intersecting bicycle facility": "Intersecting Bicycle Facility",
    "intersection approach": "Intersection Approach",
    "intersection or road crossing": "Intersection or Road Crossing",
    "crossing facility": "Crossing Facility",
    "number of lanes – adjacent road": "Number of lanes – adjacent road",
    "number of lanes - adjacent road": "Number of lanes – adjacent road",
    "number of lanes – intersecting road": "Number of lanes – intersecting road",
    "number of lanes - intersecting road": "Number of lanes – intersecting road",
    "property access": "Property Access",
    "peak pedestrian flow along or across facility": "Peak pedestrian flow along or across facility",
    "peak bicycle/lv traffic flow": "Peak bicycle/LV traffic flow",
    "observed proportion of cargo bikes and mopeds": "Observed proportion of cargo bikes and mopeds",
    "bicycle/lv speed – average": "Bicycle/LV speed – average",
    "bicycle/lv speed - average": "Bicycle/LV speed – average",
    "bicycle/lv speed differential": "Bicycle/LV speed differential",
    "road aadt": "Road AADT",
    "heavy vehicle flow": "Heavy vehicle flow",
    "road speed limit": "Road speed limit",
    "road operating speed (mean)": "Road operating speed (mean)",
    "road operating speed (mean) - km/h": "Road operating speed (mean)",
    "road operating speed (unit)": "Road operating speed (unit)",
}


# ============================ CycleRAP Model =================================

def extract_lookup_tables(ws):
    """Attribute factor blocks: CZ (category) / DA (attribute) / DC (cat id) /
    DD (option) / DE (risk factor) / DF (conditional flag), rows 3-83."""
    tables = {}
    current = None
    for r in range(3, 84):
        attr_label = ws.cell(r, ci("DA")).value
        cat_id = ws.cell(r, ci("DC")).value
        if attr_label:
            key = LABEL_TO_MODULE_KEY.get(norm_label(attr_label))
            if key is None:
                raise SystemExit(f"Unmapped model attribute at DA{r}: {attr_label!r}")
            current = key
            tables.setdefault(key, {})
        if current is None or not isinstance(cat_id, (int, float)):
            continue
        risk = ws.cell(r, ci("DE")).value
        cond = ws.cell(r, ci("DF")).value
        tables[current][int(cat_id)] = {
            "risk": float(risk) if risk is not None else 1.0,
            "cond": int(cond) if cond is not None else 0,
            "label": str(ws.cell(r, ci("DD")).value),
        }
    return tables


def extract_facility_table(ws):
    """DN3:DT8 — facility type: DP tr_vehicle_impact, DQ bp_cond, DR vb_cond,
    DS vb_sev, DT vb_cf."""
    out = {}
    for r in range(3, 9):
        cat = int(ws.cell(r, ci("DN")).value)
        out[cat] = {
            "label": str(ws.cell(r, ci("DO")).value).replace("\xa0", " ").strip(),
            "tr_vehicle_impact": float(ws.cell(r, ci("DP")).value),
            "bp_cond": int(ws.cell(r, ci("DQ")).value),
            "vb_cond": int(ws.cell(r, ci("DR")).value),
            "vb_sev": float(ws.cell(r, ci("DS")).value),
            "vb_cf": float(ws.cell(r, ci("DT")).value),
        }
    return out


def extract_aadt_table(ws):
    """DI3:DK19 — VLOOKUP(...,TRUE) approximate match: floor to threshold."""
    out = []
    for r in range(3, 20):
        thr = ws.cell(r, ci("DI")).value
        risk = ws.cell(r, ci("DK")).value
        if thr is None:
            continue
        out.append({"threshold": float(thr), "risk": float(risk)})
    return out


def extract_speed_table(ws):
    """DD93:DF243 — factor per integer speed 0..150; DE = km/h, DF = mph.

    The model looks up ROUND(speed, 0) with exact match (CycleRAP Model DE90 =
    ROUND(O58,0); DB91 OFFSET/MATCH exact), so consumers should round-half-up
    and clamp to [0, 150].
    """
    kmh, mph = {}, {}
    for r in range(93, 244):
        spd = ws.cell(r, ci("DD")).value
        if spd is None:
            continue
        kmh[int(spd)] = float(ws.cell(r, ci("DE")).value)
        mph[int(spd)] = float(ws.cell(r, ci("DF")).value)
    assert sorted(kmh) == list(range(0, 151)), "speed table must cover 0..150"
    return {
        "kmh": [kmh[s] for s in range(0, 151)],
        "mph": [mph[s] for s in range(0, 151)],
    }


def compute_baselines(lookup_tables):
    """Safest coded option per attribute: lowest risk, tie-broken by cond=0
    then lowest cat id. Facility type is pinned to 3 (off-road bicycle path),
    matching the historical baseline used for top-contributor analysis."""
    baselines = {}
    for key, opts in lookup_tables.items():
        best = min(opts.items(), key=lambda kv: (kv[1]["risk"], kv[1]["cond"], kv[0]))
        baselines[key] = best[0]
    return baselines


# ================================= STM =======================================

# STM row layout (col S attribute label / col U cat id or operator / col V input ref)
ATTR_ROWS = range(22, 111)      # categorical attribute rows (incl. 22-25 area type)
HEAVY_ROWS = range(115, 117)    # heavy vehicle flow rows
UNIT_ROWS = range(122, 124)     # road operating speed unit rows
AADT_ROWS = {111: "ge", 112: "gt", 113: "lt", 114: "le"}
SPEED_ROWS = {118: "ge", 119: "gt", 120: "lt", 121: "le"}

CATEGORICAL_ROWS = list(ATTR_ROWS) + list(HEAVY_ROWS) + list(UNIT_ROWS)

# (treatment_id, backend field) -> corrected outcome code where the workbook
# outcome formula contradicts its own description column (see module docstring).
OUTCOME_OVERRIDES = {
    (3, "Facility access"): 1,   # desc "Adequate"; formula points at G26 (=2)
    (5, "Facility access"): 1,   # same slip on the mph variant
}


def build_row_map(ws_v):
    """Map STM row -> (backend_field, cat_id) using cols S (label, with '=Sxx'
    inheritance resolved via cached values) and U (cat id)."""
    row_map = {}
    last_label = None
    for r in CATEGORICAL_ROWS:
        label = ws_v.cell(r, ci("S")).value
        if label not in (None, ""):
            last_label = label
        cat = ws_v.cell(r, ci("U")).value
        if not isinstance(cat, (int, float)):
            continue
        field = LABEL_TO_BACKEND_FIELD.get(norm_label(last_label))
        if field is None:
            raise SystemExit(f"Unmapped STM attribute at S{r}: {last_label!r}")
        row_map[r] = (field, int(cat))
    return row_map


def extract_treatments(ws_v):
    row_map = build_row_map(ws_v)

    # --- outcome (effects) table: rows 2-31, triples at D..G / I..L / N..Q ---
    treatments = {}
    for r in range(2, 32):
        tid = ws_v.cell(r, ci("B")).value
        if tid is None:
            continue
        tid = int(tid)
        name = re.sub(r"\s+", " ", str(ws_v.cell(r, ci("C")).value).replace("\xa0", " ")).strip()
        effects = {}
        requires_manual_value = False
        for c_attr, c_code in ((ci("E"), ci("F")), (ci("J"), ci("K")), (ci("O"), ci("P"))):
            attr_label = ws_v.cell(r, c_attr).value
            if attr_label in (None, ""):
                continue
            field = LABEL_TO_BACKEND_FIELD.get(norm_label(attr_label))
            if field is None:
                raise SystemExit(f"Unmapped outcome attribute at row {r}: {attr_label!r}")
            code = ws_v.cell(r, c_code).value
            if (tid, field) in OUTCOME_OVERRIDES:
                code = OUTCOME_OVERRIDES[(tid, field)]
            if code == ".":
                requires_manual_value = True
                continue
            if code is None:
                raise SystemExit(f"Missing outcome code for treatment {tid} field {field!r}")
            effects[field] = int(code) if float(code) == int(float(code)) else float(code)
        treatments[tid] = {
            "id": tid,
            "name": name,
            "unit_scope": "mph" if "(mph" in name.lower() else ("kmh" if "(km/h" in name.lower() else None),
            "requires_manual_value": requires_manual_value,
            "effects": effects,
            "trigger_sets": [],
        }

    # --- trigger-set columns: X (24) onwards while row 2 carries a CM ID ---
    c = ci("X")
    n_cols = 0
    while True:
        cm_id = ws_v.cell(2, c).value
        if cm_id is None:
            break
        cm_id = int(cm_id)
        attrs = {}
        for r in CATEGORICAL_ROWS:
            if ws_v.cell(r, c).value is True and r in row_map:
                field, cat = row_map[r]
                attrs.setdefault(field, []).append(cat)
        aadt = {op: float(ws_v.cell(r, c).value)
                for r, op in AADT_ROWS.items()
                if isinstance(ws_v.cell(r, c).value, (int, float))}
        speed = {op: float(ws_v.cell(r, c).value)
                 for r, op in SPEED_ROWS.items()
                 if isinstance(ws_v.cell(r, c).value, (int, float))}
        tset = {"attrs": attrs}
        if aadt:
            tset["aadt"] = aadt
        if speed:
            tset["speed"] = speed
        if attrs or aadt or speed:
            treatments[cm_id]["trigger_sets"].append(tset)
        n_cols += 1
        c += 1

    print(f"  scanned {n_cols} trigger-set columns")
    return [treatments[tid] for tid in sorted(treatments)]


def main():
    print(f"Reading {XLSM.name} (cached values)...")
    wb_v = openpyxl.load_workbook(XLSM, data_only=True)
    ws_model = wb_v[MODEL_SHEET]
    ws_stm = wb_v[STM_SHEET]

    lookup_tables = extract_lookup_tables(ws_model)
    model = {
        "source_workbook": XLSM.name,
        "version": "2.14",
        "lookup_tables": lookup_tables,
        "facility_type": extract_facility_table(ws_model),
        "aadt_table": extract_aadt_table(ws_model),
        "speed_table": extract_speed_table(ws_model),
        # Inclusive upper thresholds (CycleRAP Model BU4/BU9/BU18/BU29):
        # score <= t1 -> Low, <= t2 -> Medium, <= t3 -> High, else Extreme.
        "bands": {"BB": [5, 10, 20], "BP": [5, 10, 20], "SB": [5, 10, 20], "VB": [10, 25, 60]},
        "baselines": compute_baselines(lookup_tables),
    }

    treatments = {
        "source_workbook": XLSM.name,
        "version": "2.14",
        "treatments": extract_treatments(wb_v[STM_SHEET]),
    }

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    MODEL_OUT.write_text(json.dumps(model, indent=2))
    TREATMENTS_OUT.write_text(json.dumps(treatments, indent=2))
    print(f"Wrote {MODEL_OUT} ({MODEL_OUT.stat().st_size:,} bytes)")
    print(f"  lookup_tables: {len(model['lookup_tables'])} attributes")
    print(f"  aadt_table: {len(model['aadt_table'])} breakpoints")
    print(f"  speed_table: {len(model['speed_table']['kmh'])} km/h entries")
    print(f"Wrote {TREATMENTS_OUT} ({TREATMENTS_OUT.stat().st_size:,} bytes)")
    for t in treatments["treatments"]:
        print(f"  T{t['id']:>2} {t['name'][:60]:<60} sets={len(t['trigger_sets'])} "
              f"effects={list(t['effects'].items())}")


if __name__ == "__main__":
    main()
