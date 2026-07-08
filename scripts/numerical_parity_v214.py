"""Numerical parity check: Python engine vs the v2.14 workbook's cached values.

Reads the coded segment stored in the workbook itself (`CycleRAP Model` column
O, rows 15-59) plus the cached scoring outputs (BU3/BU8/BU17/BU28/BR3 and the
CM/CD intermediates), feeds the same inputs through the native engine, and
asserts exact (<=1e-9) agreement. Also verifies the speed-table and AADT-table
values in the runtime JSON against the workbook cells, and the band boundary
semantics against the BU4/BU9/BU18/BU29 formulas.

Run:  backend/venv/bin/python scripts/numerical_parity_v214.py
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))
warnings.filterwarnings("ignore")

import openpyxl  # noqa: E402
import pandas as pd  # noqa: E402
from openpyxl.utils import column_index_from_string as ci  # noqa: E402

from app.services import cyclerap_scoring as cs  # noqa: E402

XLSM = ROOT / "backend" / "CycleRAP - model - generation v2.14-to supliers.xlsm"

# CycleRAP Model input row (column O) -> backend field. Rows 15-59.
O_ROW_TO_FIELD = {
    16: cs.FACILITY_TYPE, 17: cs.FACILITY_ACCESS, 18: cs.LOOSE_SURFACE,
    19: cs.TRAM_RAILS, 20: cs.SURFACE_DEFORMATION, 21: cs.FIXED_OBSTACLE,
    22: cs.NON_FIXED_OBSTACLE, 23: cs.LINE_OF_SIGHT, 24: cs.DELINEATION,
    25: cs.LIGHT_SEGREGATION, 26: cs.FACILITY_WIDTH, 27: cs.FLOW_DIRECTION,
    28: cs.WIDTH_RESTRICTION, 29: cs.ADJACENT_ROAD_0_1M,
    30: cs.ADJACENT_PARKING_0_1M, 31: cs.ADJACENT_HAZARD_0_1M,
    32: cs.ADJACENT_OBJECT_0_1M, 33: cs.ADJACENT_SIDEWALK_0_1M,
    34: cs.ADJACENT_ROAD_1_3M, 35: cs.ADJACENT_PARKING_1_3M,
    36: cs.ADJACENT_HAZARD_1_3M, 37: cs.ADJACENT_OBJECT_1_3M,
    38: cs.ADJACENT_SIDEWALK_1_3M, 39: cs.GRADE, 40: cs.CURVATURE,
    41: cs.STREET_LIGHTING, 42: cs.PEDESTRIAN_CROSSING,
    43: cs.INTERSECTING_FACILITY, 44: cs.INTERSECTION_APPROACH,
    45: cs.INTERSECTION_CROSSING, 46: cs.CROSSING_FACILITY,
    47: cs.NUM_LANES_ADJACENT, 48: cs.NUM_LANES_INTERSECTING,
    49: cs.PROPERTY_ACCESS, 50: cs.PEDESTRIAN_FLOW, 51: cs.BICYCLE_FLOW,
    52: cs.CARGO_BIKES, 53: cs.BICYCLE_SPEED, 54: cs.SPEED_DIFFERENTIAL,
    55: cs.ROAD_AADT, 56: cs.HEAVY_VEHICLE, 58: cs.ROAD_SPEED,
    59: cs.SPEED_UNIT,
}

TOL = 1e-9
failures = []


def check(name, got, exp):
    ok = abs(got - exp) <= TOL
    status = "OK " if ok else "FAIL"
    print(f"  [{status}] {name:<28} got={got:.12f} exp={exp:.12f}")
    if not ok:
        failures.append(name)


def main():
    print(f"Reading {XLSM.name} (cached values)...")
    wb = openpyxl.load_workbook(XLSM, data_only=True)
    ws = wb["CycleRAP Model"]

    row = {field: ws.cell(r, 15).value for r, field in O_ROW_TO_FIELD.items()}
    s = pd.Series(row)

    print("Workbook segment parity:")
    cm3 = cs.calculate_cm3(s)
    cm17 = cs.calculate_cm17(s)
    cm27 = cs.calculate_cm27(s)
    cm42 = cs.calculate_cm42(s)
    check("CM3 (avoidance manoeuvre)", cm3, ws["CM3"].value)
    check("CM17 (loss of ctrl-speed)", cm17, ws["CM17"].value or 0)
    check("CM27 (loss of ctrl-func)", cm27, ws["CM27"].value or 0)
    check("CM42 (vehicle impact)", cm42, ws["CM42"].value or 0)

    bb, bp, sb, vb, total = cs.calculate_cyclerap_score(s, cm3, cm17, cm27, cm42)
    check("BB (BU3)", bb, ws["BU3"].value)
    check("BP (BU8)", bp, ws["BU8"].value)
    check("SB (BU17)", sb, ws["BU17"].value)
    check("VB (BU28)", vb, ws["BU28"].value)
    check("Total (BR3)", total, ws["BR3"].value)

    print("Speed table parity (DD93:DF243 vs runtime JSON):")
    n_bad = 0
    for r in range(93, 244):
        spd = int(ws.cell(r, ci("DD")).value)
        for col, key in ((ci("DE"), "kmh"), (ci("DF"), "mph")):
            exp = float(ws.cell(r, col).value)
            got = cs.SPEED_RISK_TABLE[key][spd]
            if abs(got - exp) > TOL:
                n_bad += 1
                print(f"  [FAIL] speed {spd} {key}: got={got} exp={exp}")
    print(f"  [{'OK ' if n_bad == 0 else 'FAIL'}] 302 speed factors compared, {n_bad} mismatches")
    if n_bad:
        failures.append("speed_table")

    print("AADT table parity (DI3:DK19 vs runtime JSON):")
    n_bad = 0
    for i, r in enumerate(range(3, 20)):
        exp_thr = float(ws.cell(r, ci("DI")).value)
        exp_risk = float(ws.cell(r, ci("DK")).value)
        entry = cs.AADT_LOOKUP_TABLE[i]
        if abs(entry["threshold"] - exp_thr) > TOL or abs(entry["risk"] - exp_risk) > TOL:
            n_bad += 1
            print(f"  [FAIL] aadt row {r}: got={entry} exp=({exp_thr}, {exp_risk})")
    print(f"  [{'OK ' if n_bad == 0 else 'FAIL'}] 17 AADT breakpoints compared, {n_bad} mismatches")
    if n_bad:
        failures.append("aadt_table")

    print("Band boundary semantics (BU4/BU29 formulas: inclusive <=):")
    cases = [
        (5.0, "BB", 1), (5.0000001, "BB", 2), (10.0, "BB", 2), (20.0, "BB", 3),
        (20.0000001, "BB", 4), (10.0, "VB", 1), (25.0, "VB", 2), (60.0, "VB", 3),
        (60.0000001, "VB", 4),
    ]
    for score_val, ct, exp_band in cases:
        got = cs.calculate_risk_band_for_type(score_val, ct)
        status = "OK " if got == exp_band else "FAIL"
        print(f"  [{status}] band({score_val}, {ct}) = {got} (exp {exp_band})")
        if got != exp_band:
            failures.append(f"band_{ct}_{score_val}")

    print()
    if failures:
        print(f"PARITY FAILED: {len(failures)} check(s): {failures}")
        sys.exit(1)
    print("PARITY OK: Python engine matches the v2.14 workbook exactly.")


if __name__ == "__main__":
    main()
