"""STM treatment parity check: Python catalog vs the v2.14 workbook's cache.

Three legs, all against the segment coded inside the workbook itself:

1. Per-trigger-set applicability — our `trigger_set_matches` vs the cached
   pass flag in STM row 9 for each of the ~107 trigger-set columns.
   (Known caveats — the workbook's own aggregate formulas contain authoring
   bugs we deliberately do NOT replicate: (a) a few per-row match flags in
   column W have off-by-N references, e.g. `W46 = U46=V44` compares the
   Delineation row against the Line-of-Sight input; (b) the "#triggers
   matched" row-6 SUMPRODUCT only spans rows 22-110 and 115-116, so a trigger
   set that marks the speed-unit rows 122/123 can never fully match in Excel
   even when every condition holds. We implement the intended semantics, so
   such differences are reported as WARNs, not failures.)
2. Treatment-level applicability — a treatment triggers if ANY set passes;
   compared against the union of its columns' row-9 flags (unit-row caveat
   above applies).
3. "STM Results" sheet — every treatment the workbook lists as triggered must
   be applicable per the Python catalog, and the sheet's cached score columns
   (which mirror the segment's PRE-treatment model outputs — the macro pastes
   post-treatment values only at run time) must equal the native engine's
   before-scores.

Run:  backend/venv/bin/python scripts/stm_parity_v214.py
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))
warnings.filterwarnings("ignore")

import openpyxl  # noqa: E402
import pandas as pd  # noqa: E402
from openpyxl.utils import column_index_from_string as ci, get_column_letter  # noqa: E402

from app.services import cyclerap_scoring as cs  # noqa: E402
from app.services.cyclerap_scoring import calculate_cyclerap_score_native  # noqa: E402
from app.services.treatment_catalog import (  # noqa: E402
    TREATMENTS,
    TREATMENT_BY_ID,
    is_treatment_applicable,
    trigger_set_matches,
)
from numerical_parity_v214 import O_ROW_TO_FIELD  # noqa: E402

XLSM = ROOT / "backend" / "CycleRAP - model - generation v2.14-to supliers.xlsm"
TOL = 1e-9
failures = []


def main():
    print(f"Reading {XLSM.name} (cached values)...")
    wb = openpyxl.load_workbook(XLSM, data_only=True)
    ws_model = wb["CycleRAP Model"]
    ws_stm = wb["STM"]

    row = {field: ws_model.cell(r, 15).value for r, field in O_ROW_TO_FIELD.items()}

    # ---- leg 1 + 2: trigger-set / treatment applicability ------------------
    print("Trigger-set applicability vs STM row 9:")
    wb_triggered: dict[int, bool] = {}
    py_triggered: dict[int, bool] = {}
    mismatches = 0
    c = ci("X")
    set_idx: dict[int, int] = {}
    while ws_stm.cell(2, c).value is not None:
        tid = int(ws_stm.cell(2, c).value)
        set_idx[tid] = set_idx.get(tid, 0)
        excel_flag = bool(ws_stm.cell(9, c).value)
        t = TREATMENT_BY_ID[tid]
        if set_idx[tid] < len(t["trigger_sets"]):
            py_flag = trigger_set_matches(row, t["trigger_sets"][set_idx[tid]])
        else:
            py_flag = False  # empty column dropped at extraction
        wb_triggered[tid] = wb_triggered.get(tid, False) or excel_flag
        py_triggered[tid] = py_triggered.get(tid, False) or py_flag
        if py_flag != excel_flag:
            mismatches += 1
            tset = t["trigger_sets"][set_idx[tid]] if set_idx[tid] < len(t["trigger_sets"]) else {}
            unit_marked = "Road operating speed (unit)" in tset.get("attrs", {})
            reason = " (workbook unit-row counting bug)" if unit_marked and py_flag else ""
            print(f"  [WARN] T{tid} set {set_idx[tid] + 1} (col {get_column_letter(c)}): "
                  f"python={py_flag} workbook={excel_flag}{reason}")
        set_idx[tid] += 1
        c += 1
    print(f"  compared {c - ci('X')} trigger-set columns, {mismatches} set-level differences")

    print("Treatment-level applicability:")
    t_mismatch = 0
    for t in TREATMENTS:
        tid = t["id"]
        py = is_treatment_applicable(row, t)
        exc = wb_triggered.get(tid, False)
        # python-True / workbook-False is expected when every passing set
        # marks the speed-unit row (workbook counting bug, see docstring).
        unit_bug = py and not exc and all(
            "Road operating speed (unit)" in ts.get("attrs", {})
            for ts in t["trigger_sets"] if trigger_set_matches(row, ts)
        )
        mark = "OK " if py == exc else ("WARN" if unit_bug else "FAIL")
        if py != exc and not unit_bug:
            t_mismatch += 1
            failures.append(f"T{tid}_applicability")
        if py or exc or py != exc:
            print(f"  [{mark}] T{tid:>2} {t['name'][:58]:<58} python={py} workbook={exc}")
    print(f"  {t_mismatch} treatment-level mismatches (excluding unit-row-bug WARNs)")

    # ---- leg 3: STM Results sheet ------------------------------------------
    print("'STM Results' sheet (triggered list + pre-treatment score mirror):")
    ws_res = wb["STM Results"]
    before_df = calculate_cyclerap_score_native(pd.DataFrame([row]))
    for r in range(2, ws_res.max_row + 1):
        tid = ws_res.cell(r, 1).value
        if tid is None:
            continue
        tid = int(tid)
        py = is_treatment_applicable(row, TREATMENT_BY_ID[tid])
        ok = py is True
        if not ok:
            failures.append(f"T{tid}_listed_but_not_applicable")
        print(f"  [{'OK ' if ok else 'FAIL'}] T{tid:>2} listed in STM Results; python applicable={py}")
        for col, key in ((11, "BB"), (12, "BP"), (13, "SB"), (14, "VB"), (15, "Overall Risk Level")):
            exp = float(ws_res.cell(r, col).value)
            got = float(before_df[key].iloc[0])
            if abs(got - exp) > 1e-4:  # engine rounds to 4 dp
                failures.append(f"T{tid}_prescore_{key}")
                print(f"  [FAIL] T{tid:>2} pre-score {key}: got={got:.4f} exp={exp:.4f}")

    print()
    if failures:
        print(f"STM PARITY FAILED: {failures}")
        sys.exit(1)
    print("STM PARITY OK (trigger applicability + STM Results list/scores match).")


if __name__ == "__main__":
    main()
